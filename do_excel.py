# -*- coding: UTF-8 -*-
# @time     : 2021-03-15 16:46
# @Auther   : Aaron
# @File     : do_excel.py

import os

from openpyxl import load_workbook


class TestCase:
    pass


class DoExcel:

    def __init__(self, filename, sheetname=None):
        self.filename = filename
        self.sheetname = sheetname

    def read_data(self):

        wb = load_workbook(self.filename)
        if self.sheetname is None:
            ws = wb.active
        else:
            ws = wb[self.sheetname]
        testcases_list = []
        header_list = []
        for row in range(1, ws.max_row+1):
            one_testcase = TestCase()
            for column in range(1, ws.max_column+1):
                one_cell_value = ws.cell(row, column).value

                if row == 1:
                    header_list.append(one_cell_value)
                else:
                    key = header_list[column-1]
                    if key == "actual":
                        setattr(one_testcase, "actual_column", column)
                    elif key == "result":
                        setattr(one_testcase, "result_column", column)

                    setattr(one_testcase, key, one_cell_value)

            if row != 1:
                setattr(one_testcase, "row", row)
                testcases_list.append(one_testcase)

        return testcases_list

    def write_data(self, one_testcase, actual_value, result_value):
        wb = load_workbook(self.filename)
        if self.sheetname is None:
            ws = wb.active
        else:
            ws = wb[self.sheetname]

        ws.cell(one_testcase.row, one_testcase.actual_column, value=actual_value)
        ws.cell(one_testcase.row, one_testcase.result_column, value=result_value)
        wb.save(self.filename)


if __name__ == '__main__':
    excel_filename = "create_user.xlsx"
    sheet_name = "create_user"
    doexcel = DoExcel(excel_filename, sheet_name)
    doexcel.read_data()


